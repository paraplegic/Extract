#!/usr/bin/env python

import sys
import openpyxl 
import datetime

##
## replace multiple spaces in string with a single one.
def single_space( s ):
	return ' '.join(s.split())

##
## return a cell value as a string ...
def get_cell_value( cell ):

	v = cell.value
	if not v:
		return ''

	if type( v ) == str:
		return single_space( v )

	if type( v ) == float:
		rv = "%10.2f" % v
		return rv.strip()

	if type( v ) == long:
		return "%d" % v

	if type(v) == unicode:
		return single_space( v.encode('utf-8') )

	if type( v ) == datetime.datetime:
		return v.strftime( "%Y.%m.%d_%H:%M:%S" )

	return single_space( v )

##
## return a boolean true if row is <empty> ...
def empty_row( row ):
    for v in row:
        if v.value is not None:
            return False

    return True

def no_values( r ):
    tl = 0
    for v in r:
        tl += len( v )

    if tl == 0:
        return True

    return False


##
## output with some control ...
def out( s ):
    if s:
        sys.stdout.write( s )

##
## return true if token contains digits 
def token_has_digits( instr ):
    return any(char.isdigit() for char in instr)

def styleList( r ):
    rv = []
    r[0] = r[0].replace( " CLG", "CLG" )
    for s in r:
        tl = s.split()
        for tkn in tl:
            if "100%" in tkn or "STYLE:" in tkn:
                continue
            if token_has_digits( tkn ):
                rv.append( tkn )

    return unique_list( rv )

##
## recode a list of cells to a list of values ...
def recode( row ):
	rv=[]
	for x in row:
		c_val = get_cell_value( x )
		if c_val:
			rv.append( c_val )
		else: 
			rv.append( '' )
	return rv

def create_maps( styles, hdr0, hdr1 ):
    rv = []
    for s in styles:
        rv.append( create_map( styles, hdr0, hdr1 ) )
    return rv

def create_map( styles, hdr0, hdr1 ):
	pl = [] 
	dim = [] 
	wgt = [] 
	cube = None
	for pt in price_tags:
		if pt in hdr0:
			pl.append( hdr0.index( pt ) )

	dim = [i for i, elem in enumerate(hdr0) if 'DIMENSIONS' in elem]
	grades = range( pl[0], dim[0] )

        leather = False
        if "100% all LEATHER" in hdr1[0]:
            leather = True

	for w in weight_tags:
		if w in hdr1:
			wgt.append( hdr1.index( w ) )
		
        rv = { "styles": styles, "price_lists": pl, "dimension": dim, "weight": wgt, "grades": grades, "hdr1": hdr1, "leather": leather }
##	out( "%s\n" % rv )
	return rv

def unique_list( l ):
    rv = set( l )
    return list( rv )

def parse_row( r ):
    des = r[0]
    des = des.replace( ' ___ ', '' )
    dl = des.split( ' ' )
    style = dl[0]
    if '-' in style:
        style = style.replace( '-', ' ' )
        xx = style.split()
        style = xx[0]
        model = xx[1].replace( '-', '' )
        desc = " ".join( dl[1:] ).replace( '-', '' ).strip()
    else:
        model = " ".join( dl[1:] ).replace( '-', '' )
        desc = " ".join( dl[2:] ).replace( '-', '' ).strip()

    if style.endswith( "Q" ) or style.endswith( "K" ):
        model = style[-1]
        style = style[0:-1]

    if style == "100%":
        return ( None, None, None )

    return ( style, model, desc )

def walk_map( map, row ):
    (s,m,d) = parse_row( row )
    prv_grade = 0
    prv_price = 0
    grade = 0
    price = 0
    sgrade = ""
    for g in map['grades']:
        if row[g]:
            if s and m and d:
                sgrade = map['hdr1'][g]
                sprice = row[g]
                prv_price = price
                prv_grade = grade
                if token_has_digits( sgrade ):
                    grade = int( sgrade )
                    price = int( row[g] )
                out( "%s," % s )
                out( "%s," % m )
                out( "%s," % d )
                out( "%s," % sgrade )
                out( "%s\n" % sprice  )

    delta = grade - prv_grade
    bump = price - prv_price
    new = price
    if not map['leather'] and sgrade != "MML" and delta != 0:
        for x in range( grade+delta, 149, delta ):
            out( "%s," % s )
            out( "%s," % m )
            out( "%s," % d )
            new += bump
            out( "%s," % x )
            out( "%s\n" % new  )

price_tags = [
	"MSRP",
	"CDN PRICING",
	"CDN FABRIC PRICING",
	"CDN LEATHER PRICING",
	"MSRP FABRIC PRICING",
	"MSRP LEATHER PRICING"
]

weight_tags = [
	"(KG)",
	"(LBS)"
]

stop_tags = [
	"DIMENSIONS",
]

otto_fstyle = 'Fabric Style'
otto_lstyle = 'Leather Style'
otto_fabric = 'CDN FABRIC PRICING'
otto_leather = 'CDN LEATHER PRICING'
otto_end = 'DIMENSIONS'
def map_ottoman( hdr0, hdr1 ):

    fstyle = hdr0.index( otto_fstyle )
    lstyle = hdr0.index( otto_lstyle )
    fstart = hdr0.index( otto_fabric )
    lstart = hdr0.index( otto_leather )
    lend = hdr0.index( otto_end )

    fprices = range( fstart, lstart )
    lprices = range( lstart, lend )

    rv = { "fstyle": fstyle, "lstyle": lstyle, "fgrades": fprices, "lgrades": lprices, "hdr1": hdr1 } 
    return rv

def ottoman( map, r ):
    return

    ix = map['fstyle']
    if r[ix]:
        xx = r[ix].split( ',' )
        for x in xx:
            z = x.split( '-' )
            style = z[0]
            model = z[1]



##
## main logic ...  
def main( args ):

	fn = args[1]
	wb = openpyxl.load_workbook( fn, data_only = True, read_only = True )
	for sn in wb.sheetnames:
                if "ottoman" in sn.lower():
                    continue

                if not "wholesale" in sn.lower():
                    continue

		nhdrs = 0
		last_state = None
		state = "Searching"
		for r in wb[sn].rows:

			if empty_row( r ):
				continue

			v = recode( r )
                        if no_values( v ):
                            continue

			s = ",".join( v )
			if "STYLE: " in v[0]:
				state = "Found Style"
                                styles = styleList( v )
##				print state, v[0], styleList( v )
				continue

			if "DESCRIPTION" == v[0]:
				state = "Found Description"
				nhdrs += 1
##				print state, nhdrs
				h0 = v
				continue

			if state == "Found Description":
				state = "Found Header"
				maps = create_maps( styles, h0, v )
				state = "Expect Data"
				continue

			if state == "Expect Data":
                            for map in maps:
				walk_map( map, v )
		else:
			last_state = state
			if not last_state == "Found Style":
				state = "Searching"

        h0 = None
	for sn in wb.sheetnames:
            if not "wholesale ottomans" in sn.lower():
                continue

            last_state = None
            state = "Searching"
            for r in wb[sn].rows:

                    if empty_row( r ):
                        continue

                    v = recode( r )
                    if no_values( v ):
                        continue

                    if "DESCRIPTION" == v[0]:
                        state = "Found Description"
                        nhdrs += 1
                        h0 = v
                        continue

                    if state == "Found Description":
                        state = "Found Header"
                        map = map_ottoman( h0, v )
                        state = "Expect Data"
                        continue

                    if state == "Expect Data":
                        ottoman( map, v )


## a magic incantation ...
if __name__ == '__main__':
    main( sys.argv )
